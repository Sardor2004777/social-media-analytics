"""Excel report generator (openpyxl).

Produces a multi-sheet XLSX for a given user:

    Summary    — 6 KPI rows, one row per platform.
    Posts      — all posts with metrics.
    Comments   — all comments with sentiment + language.
    Sentiment  — label counts per language.
    Platforms  — account-level stats with a pie chart.

The workbook is returned as raw bytes so the view can stream it with the right
Content-Type / Content-Disposition without touching disk.
"""
from __future__ import annotations

import io
from collections import Counter
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.db.models import Avg, Count, Sum
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from apps.analytics.models import SentimentResult
from apps.collectors.models import Comment
from apps.social.models import ConnectedAccount, Post

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="4F46E5")  # brand-600
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill("solid", fgColor="EEF2FF")  # brand-50
SUBHEAD_FONT = Font(bold=True, color="3730A3", size=11)
BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def _apply_header(ws: Worksheet, row: int, values: Iterable) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 22


def _autosize(ws: Worksheet, widths: dict[str, float] | None = None) -> None:
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        return
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(ws: Worksheet, user) -> None:
    ws.title = "Summary"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Social Analytics Report"
    ws["A1"].font = Font(bold=True, color="1E1B4B", size=18)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Foydalanuvchi: {user.email}"
    ws["A2"].font = Font(italic=True, color="475569")
    ws["A3"] = f"Sana: {datetime.now():%Y-%m-%d %H:%M}"
    ws["A3"].font = Font(italic=True, color="475569")

    accounts = ConnectedAccount.objects.filter(user=user)
    posts = Post.objects.filter(account__user=user)
    comments = Comment.objects.filter(post__account__user=user)
    sentiments = SentimentResult.objects.filter(comment__post__account__user=user)

    avg_eng_val = posts.aggregate(v=Avg("engagement_rate"))["v"] or 0
    total_likes = posts.aggregate(v=Sum("likes"))["v"] or 0
    total_views = posts.aggregate(v=Sum("views"))["v"] or 0

    pos = sentiments.filter(label="positive").count()
    neu = sentiments.filter(label="neutral").count()
    neg = sentiments.filter(label="negative").count()
    total_s = pos + neu + neg or 1

    kpis = [
        ("Akkauntlar",           accounts.count()),
        ("Jami postlar",         posts.count()),
        ("Jami kommentlar",      comments.count()),
        ("Jami layklar",         total_likes),
        ("Jami ko'rishlar",      total_views),
        ("O'rtacha engagement",  f"{avg_eng_val*100:.2f}%"),
        ("Pozitiv sentiment",    f"{pos*100/total_s:.1f}%"),
        ("Neytral sentiment",    f"{neu*100/total_s:.1f}%"),
        ("Negativ sentiment",    f"{neg*100/total_s:.1f}%"),
    ]

    _apply_header(ws, 5, ("Ko'rsatkich", "Qiymat"))
    for i, (label, value) in enumerate(kpis, start=6):
        ws.cell(row=i, column=1, value=label).border = BORDER
        c = ws.cell(row=i, column=2, value=value)
        c.border = BORDER
        c.font = Font(bold=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20


def _build_posts(ws: Worksheet, user) -> None:
    ws.title = "Posts"
    headers = ("Platforma", "Handle", "Turi", "Caption", "Postlar sanasi",
               "Ko'rishlar", "Layklar", "Kommentlar", "Shares", "Engagement %")
    _apply_header(ws, 1, headers)

    qs = (
        Post.objects.filter(account__user=user)
        .select_related("account")
        .order_by("-published_at")
    )
    for r, p in enumerate(qs, start=2):
        ws.cell(row=r, column=1, value=p.account.platform)
        ws.cell(row=r, column=2, value=p.account.handle)
        ws.cell(row=r, column=3, value=p.post_type)
        ws.cell(row=r, column=4, value=(p.caption or "")[:180])
        c5 = ws.cell(row=r, column=5, value=p.published_at.replace(tzinfo=None))
        c5.number_format = "yyyy-mm-dd hh:mm"
        c6 = ws.cell(row=r, column=6, value=p.views)
        c6.number_format = "#,##0"
        c7 = ws.cell(row=r, column=7, value=p.likes)
        c7.number_format = "#,##0"
        c8 = ws.cell(row=r, column=8, value=p.comments_count)
        c8.number_format = "#,##0"
        c9 = ws.cell(row=r, column=9, value=p.shares)
        c9.number_format = "#,##0"
        c10 = ws.cell(row=r, column=10, value=p.engagement_rate)
        c10.number_format = "0.00%"

    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    ws.freeze_panes = "A2"
    _autosize(ws, {"A": 12, "B": 18, "C": 14, "D": 60, "E": 20, "F": 12, "G": 12, "H": 12, "I": 10, "J": 14})


def _build_comments(ws: Worksheet, user) -> None:
    ws.title = "Comments"
    headers = ("Sana", "Platforma", "Post caption", "Muallif", "Til", "Sentiment", "Score", "Matn")
    _apply_header(ws, 1, headers)

    qs = (
        Comment.objects.filter(post__account__user=user)
        .select_related("post__account", "sentiment")
        .order_by("-published_at")[:2000]  # cap for Excel responsiveness
    )
    for r, c in enumerate(qs, start=2):
        ws.cell(row=r, column=1, value=c.published_at.replace(tzinfo=None)).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=r, column=2, value=c.post.account.platform)
        ws.cell(row=r, column=3, value=(c.post.caption or "")[:80])
        ws.cell(row=r, column=4, value=c.author_handle)
        ws.cell(row=r, column=5, value=c.language)
        sent = getattr(c, "sentiment", None)
        ws.cell(row=r, column=6, value=sent.label if sent else "")
        sc = ws.cell(row=r, column=7, value=sent.score if sent else None)
        if sent:
            sc.number_format = "0.00"
        ws.cell(row=r, column=8, value=c.body[:500])

    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    ws.freeze_panes = "A2"
    _autosize(ws, {"A": 20, "B": 12, "C": 40, "D": 20, "E": 6, "F": 12, "G": 8, "H": 70})


def _build_sentiment(ws: Worksheet, user) -> None:
    ws.title = "Sentiment"
    _apply_header(ws, 1, ("Til", "Pozitiv", "Neytral", "Negativ", "Jami"))
    qs = (
        SentimentResult.objects.filter(comment__post__account__user=user)
        .values_list("comment__language", "label")
    )
    counter: dict[tuple[str, str], int] = Counter(qs)
    languages = sorted({lang for (lang, _) in counter})
    for r, lang in enumerate(languages, start=2):
        pos = counter.get((lang, "positive"), 0)
        neu = counter.get((lang, "neutral"), 0)
        neg = counter.get((lang, "negative"), 0)
        total = pos + neu + neg
        ws.cell(row=r, column=1, value=lang or "—").border = BORDER
        for col, val in enumerate((pos, neu, neg, total), start=2):
            c = ws.cell(row=r, column=col, value=val)
            c.number_format = "#,##0"
            c.border = BORDER

    # Add a BarChart for quick visual
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "Sentiment · til bo'yicha"
    chart.y_axis.title = "Kommentlar soni"
    chart.x_axis.title = "Til"
    data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 18
    ws.add_chart(chart, "G2")

    _autosize(ws)


def _build_platforms(ws: Worksheet, user) -> None:
    ws.title = "Platforms"
    _apply_header(ws, 1, ("Platforma", "Handle", "Obunachilar", "Postlar", "Jami layk", "Engagement %"))

    rows = (
        ConnectedAccount.objects.filter(user=user)
        .annotate(
            total_posts=Count("posts"),
            total_likes=Sum("posts__likes"),
            avg_eng=Avg("posts__engagement_rate"),
        )
        .order_by("platform")
    )
    for r, a in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=a.platform).border = BORDER
        ws.cell(row=r, column=2, value=a.handle).border = BORDER
        c3 = ws.cell(row=r, column=3, value=a.follower_count)
        c3.number_format = "#,##0"
        c3.border = BORDER
        c4 = ws.cell(row=r, column=4, value=a.total_posts or 0)
        c4.number_format = "#,##0"
        c4.border = BORDER
        c5 = ws.cell(row=r, column=5, value=a.total_likes or 0)
        c5.number_format = "#,##0"
        c5.border = BORDER
        c6 = ws.cell(row=r, column=6, value=a.avg_eng or 0)
        c6.number_format = "0.00%"
        c6.border = BORDER

    # Pie chart of followers per platform
    if ws.max_row > 1:
        chart = PieChart()
        chart.title = "Obunachilar ulushi"
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)  # includes header as title
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 9
        chart.width = 14
        ws.add_chart(chart, "H2")

    _autosize(ws, {"A": 14, "B": 22, "C": 14, "D": 10, "E": 14, "F": 14})


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_workbook(user) -> bytes:
    """Return an XLSX workbook serialized to bytes for the given user."""
    wb = Workbook()

    # default active sheet becomes Summary
    _build_summary(wb.active, user)
    _build_posts(wb.create_sheet(), user)
    _build_comments(wb.create_sheet(), user)
    _build_sentiment(wb.create_sheet(), user)
    _build_platforms(wb.create_sheet(), user)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
